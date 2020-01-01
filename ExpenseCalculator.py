#!/usr/bin/env python3
import PySimpleGUI as sg
import os
from openpyxl import Workbook, load_workbook
import datetime

########################################################################################################################
#   Program calculates and records all personal expenses in a seperate
#   spreadsheet using a GUI frontend. Each expense and piece of data is stored
#   new cells while the balance is a constantly updating cell
#   Every transactions results in a new row on the spreadsheet with:
#    - a timestamp of the time the transaction was entered into GUI
#    - a record of the amount deposited or withdrawn
#    - a description of the transaction
########################################################################################################################
FILENAME = "ExpenseSheet.xlsx"

#   Workbook
if os.path.exists(FILENAME):
    print('Found existing expense sheet')
    workbook = load_workbook(FILENAME)
    sheet = workbook.active

else:
    print("Didn't find existing expense sheet.\n\nCreating new sheet")
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Balance"
    sheet['B1'] = "Timestamp"
    sheet['C1'] = "Transaction Amount"
    sheet['D1'] = "Description"
    sheet['A2'] = 0.0
    workbook.save(os.path.join(FILENAME))
    print("Success")

"""
Layout of GUI

[       Textfield('amount')  |   Balance('balance')         ]
[       Deposit Radio('deposit') Withdraw Radio('withdraw') ]
[       Description of Transaction('description') --------->]
[       {Finish Button('finish')}   {Quit Button}           ]

"""
layout = [[sg.InputText(size=(20, 1), key='amount'), sg.VerticalSeparator(),
           sg.Text(("$ " + str(sheet['A2'].value)), auto_size_text=True, key='balance')],
          [sg.Radio('Deposit', "RADIO1", default=False, auto_size_text=True, key='deposit'),
           sg.Radio('Withdraw', "RADIO1", default=False, auto_size_text=True, key='withdraw')],
          [sg.InputText('Description of transaction', size=(45, 5), key='description')],
          [sg.Button('Finish', key='finish'), sg.Button('History', key="history"), sg.Quit()]]

window = sg.Window('Expense Calculator', layout, location=(600, 550), size=(450, 150))

#   Functions
"""
    Pre:    element 'Finish' exists, and at least one of the radio elements is selected
    Post:   the deposit() or withdraw() functions are called based on user selection
"""
def calculate():
    if window.Element('deposit').Get():
        deposit()

    elif window.Element('withdraw').Get():
        withdraw()

    else:
        sg.PopupError("Please select either Deposit or Withdraw", location=(500, 500), no_titlebar=True, line_width=25)


"""
    Pre:    window element 'amount' exists
    Post:   outputs True if input is an integer/float value, False otherwise
        Needed to make sure whether or not string returned by the text elements
        are numbers, not alphanumeric characters
"""
def check_valid_amount(input):
    #   if the text field is empty, return False
    if input == "":
        return False

    #   if the text field is somehow completely void, return False
    if input is None:
        return False

    try:
        float(input)
        return True

    except ValueError:
        return False


"""
    Pre:    window element 'description' exists
    Post:   outputs True if input is different than the default text and is not empty, False otherwise
"""
def check_valid_input(input):
    #   if text field is same as default text, return false
    if input == "Description of transaction":
        return False

    if input is None:
        return False

    return True


"""
    Pre:    window elements 'deposit', 'amount', and 'description' exist and are valid inputs
    Post:   calls on the update() function to update the record with the necisary info from the transaction
"""
def deposit():
    transaction = window.Element("amount").Get()  # user entered amount
    description = window.Element("description").Get()  # user entered description
    is_valid_amount = check_valid_amount(transaction)
    is_valid_input = check_valid_input(description)

    #   loop ensures user enters text for a description
    if is_valid_input is False:
        while is_valid_input is False:
            description = sg.PopupScrolled("Please enter a valid description",
                                           size=(50, 20),
                                           location=(700, 200))
            is_valid_input = check_valid_input(description)

    #   loop ensures text entered is a number, not a string or char
    if is_valid_amount is False:
        while is_valid_amount is False:
            transaction = sg.PopupGetText("Please enter a dollar amount into the text field", "",
                                          no_titlebar=True,
                                          location=(600, 720))
            is_valid_amount = check_valid_input(transaction)

    transaction = float(transaction)  # change to float for ease
    old_balance = float(sheet['A2'].value)
    new_balance = old_balance + transaction

    update(transaction, new_balance, str(window.Element("description").Get()), "+")  # update ExpenseSheet.xlsx


"""
    Pre:    window elements 'withdraw', 'amount', and 'description' exist and are valid inputs
    Post:   calls on the update() function to update the record with the necessary info from the transaction
"""
def withdraw():
    transaction = window.Element("amount").Get()  # user entered amount
    description = window.Element("description").Get()
    is_valid_amount = check_valid_input(transaction)
    is_valid_input = check_valid_input(description)

    #   loop ensures text entered in description field is not the default value, or a illegal phrase
    if is_valid_input is False:
        while is_valid_input is False:
            transaction = sg.PopupGetText("Please enter a valid description", "",
                                          no_titlebar=True,
                                          location=(600, 700))
            is_valid_input = check_valid_input(description)

    #   loop ensures text entered is a number, not a string or char
    if is_valid_amount is False:
        while is_valid_amount is False:
            transaction = sg.PopupGetText("Please enter a dollar amount into the text field", "",
                                          no_titlebar=True,
                                          location=(600, 720))
            is_valid_amount = check_valid_input(transaction)

    transaction = float(transaction)
    old_balance = float(sheet['A2'].value)
    new_balance = old_balance - transaction

    update(transaction, new_balance, str(window.Element("description").Get()), "-")  # update ExpenseSheet.xlsx


"""
    Pre:
    Post:
"""
def create_text(prices, times, descs):
    text = []
    counter = 0

    while counter < 10:
        try:
            price = prices[counter]
            time = times[counter]
            desc = descs[counter]

            newline = str(price+'\t\t'+str(time)+'\t\t'+desc+'\n')
            text.append(newline)
            counter += 1

        except IndexError:
            break

    return text


"""
    Pre:
    Post:
"""
def show_history():
    price_arr = []
    time_arr = []
    desc_arr = []
    timestamp_col = sheet['B']
    transaction_col = sheet['C']
    description_col = sheet['D']
    counter = 0

    #   data collection
    for cells in transaction_col:
        price_arr.append(cells.value)

        if counter == 10 or cells.value is None:
            counter = 0
            break

    for cells in timestamp_col:
        time_arr.append(cells.value)

        if counter == 10 or cells.value is None:
            counter = 0
            break

    for cells in description_col:
        desc_arr.append(cells.value)

        if counter == 10 or cells.value is None:
            counter = 0
            break

    #   display new window with history
    message = create_text(price_arr, time_arr, desc_arr)
    sg.PopupScrolled(message, title="History", location=(700, 200))


"""
    Pre:    workbook ExpenseSheet.xlsx exists
    Post:   ExpenseSheet cells are updated with the new balance and transaction information is saved
"""
def update(transaction_amount, balance_update, description, sign):
    curtime = datetime.datetime.now()
    sheet['A2'].value = str(transaction_amount + balance_update)        # updates total balance
    timestamp_col = sheet['B']                                          # row for each time stamp
    transaction_col = sheet['C']                                        # row for each transaction
    description_col = sheet['D']                                        # row for each transaction description

    for cells in timestamp_col:
        if cells.value is None:
            sheet.cell(cells.row, cells.column).value = curtime
            break

    for cells in transaction_col:
        if cells.value is None:
            sheet.cell(cells.row, cells.column).value = str(sign) + str(transaction_amount)
            break

    for cells in description_col:
        if cells.value is None:
            sheet.cell(cells.row, cells.column).value = str(description)
            break

    workbook.save(os.path.join(FILENAME))
    window.Element("balance").Update("$ " + str(sheet['A2'].value))     # updates balance in GUI menu
    sg.PopupOK('Complete', location=(600, 720), no_titlebar=True)


#   event loop
while True:
    event, value = window.Read()

    if event == 'finish':
        calculate()

    if event == 'history':
        show_history()

    elif event == 'Quit' or event is None:
        window.Close()
        break
